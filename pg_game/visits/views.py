from django.shortcuts import render
from .models import LoginTime
from histories.models import History
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill
from django.db.models import Count, Q


@login_required(login_url='login')
def agents_login_list(request):
    agents_login_obj = LoginTime.objects.raw(
        'SELECT history.id, history.full_name, history.sa_code, count(login.sa_code) '
        'FROM agents_login_time as login JOIN agents_history as history '
        'on login.sa_code = history.sa_code group by history.id, history.full_name, '
        'history.sa_code')
    data = {'agents': agents_login_obj}
    return render(request, 'visit/list.html', context=data)


@login_required(login_url='login')
def agent_login_view(request, agent_id):
    agent = get_object_or_404(History, id=agent_id)
    connection_times_obj = LoginTime.objects.filter(sa_code=agent.sa_code)
    data = {'connection_times': connection_times_obj, 'agent': agent}
    return render(request, 'visit/view.html', context=data)


@login_required(login_url='login')
def login_report_page(request):
    return render(request, 'visit/report.html')


@login_required(login_url='login')
def login_report_export(request):
    if request.POST:
        date = request.POST.get('date')
        agents_obj = History.objects.all()
        workbook = Workbook()
        wb = workbook['Sheet']
        wb['A1'] = 'Дата посещения'
        wb['B1'] = 'Город'
        wb['C1'] = 'ФИО'
        wb['D1'] = 'RT код'
        wb['E1'] = 'Количество посещений'
        wb.column_dimensions['A'].widthw = wb.column_dimensions['E'].width = wb.column_dimensions['C'].width = 35
        yellow_color = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        wb['A1'].fill = wb['B1'].fill = wb['C1'].fill = wb['D1'].fill = wb['E1'].fill = yellow_color
        excel_column = 2
        for agent in agents_obj:
            agent_filter = {'sa_code': agent.sa_code}
            if date:
                agent_filter['date_time__date'] = date
            agents_login_dict = LoginTime.objects.filter(create_filter_query(agent_filter)).extra(
                select={'date_time': 'date( agents_login_time.date_time )'}).values('sa_code', 'date_time').annotate(
                login_count=Count('date_time'))
            for agent_login in agents_login_dict:
                wb['A' + str(excel_column)] = agent_login['date_time']
                wb['B' + str(excel_column)] = agent.branch_name
                wb['C' + str(excel_column)] = agent.full_name
                wb['D' + str(excel_column)] = agent.sa_code
                wb['E' + str(excel_column)] = agent_login['login_count']
                excel_column += 1
        response = HttpResponse(content=save_virtual_workbook(workbook),
                                content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        return response
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


def create_filter_query(filter_dict):
    filters = Q()
    for item in filter_dict:
        filters &= Q(**{item: filter_dict[item]})
    return filters
